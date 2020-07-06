#!/usr/bin/python
'''pixelArt.py - Real images to Excel sheets
Leonardo Dias - 12/05/2020
'''

from sys                import argv
from PIL                import Image
from pathlib            import Path
from openpyxl           import Workbook
from os                 import chdir
from sys                import argv
from datetime           import datetime
from openpyxl.utils     import get_column_letter
from openpyxl.styles    import PatternFill, fills

def convt(dec: int) -> str:
    'Converter [0, 255] -> [00, FF]'
    hexSyst = (lambda i: ['0', '1', '2', '3', '4', '5', '6', '7',
                          '8', '9', 'A', 'B', 'C', 'D', 'E', 'F'][i])
    return hexSyst(dec//16)+hexSyst(dec%16)

def main(path):

    wb = Workbook()
    s1 = wb.active
    im = Image.open(path.name)

    print('\nCriando imagem pixelada...')

    width, height = im.size

    # Encontrar maior múltiplo width, height
    redu = 1
    for i in range(2, int(max(width, height)**(1/2))+1):
        if width%i==0 and height%i==0: redu = i

    pixels = {}
    for x in range(0, width, redu):
        for y in range(0, height, redu):
            r, g, b = 0, 0, 0
            for i in range(redu):
                r1, g1, b1 = im.getpixel((x+i, y+i))
                r += r1
                g += g1
                b += b1
            pixels.setdefault((x, y), (0, 0, 0))
            for i in range(redu):
                for j in range(redu):
                    pixels[(x+i, y+j)] = (r//redu, g//redu, b//redu)

    newIm = Image.new('RGB', (width, height))
    for x in range(width):
        for y in range(height):
            newIm.putpixel((x, y), pixels[(x, y)])
    newIm.save(f'{path.name}_{datetime.now().date()}.png')
    print('Imagem criada: ',f'{path.name}_{datetime.now().date()}.png\n\nCriando imagem pixelada...')




    Gimp = newIm.resize((int(width/redu), int(height/redu)))
    width, height = Gimp.size
    if max(width, height)>150:
        amp = 72/max(width, height)
        Gimp = Gimp.resize((int(width*amp), int(height*amp)))
        width, height = Gimp.size

    Gimp.save(f'Pixels_{path.name}_{datetime.now().date()}.png')
    print('Imagem pixelada para GIMP criada\n\nCriando agora arquivo Excel...')

    excelIm = Gimp.copy()
    width, height = excelIm.size

    for x in range(width):
        for y in range(height):
            r, g, b = excelIm.getpixel((x, y))
            r = convt(r)
            g = convt(g)
            b = convt(b)
            color = f'{r}{g}{b}'
            position = f'{get_column_letter(x+1)}{y}'
            s1.row_dimensions[y].height = 10
            s1.column_dimensions[get_column_letter(x+1)].width = 2
            try: s1[position].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            except AttributeError: s1[position][0].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    wb.save(f'{path.name}_{datetime.now().date()}.xlsx')

    print('\n\nImagem criada com sucesso :)\n\n')

if __name__=='__main__':

    print('\n', '-'*60, '\n', 'Programa Pixelador de panilha para PixelArt'.center(60, ' '),
          '\n', 'Autor: Leonardo Dias'.center(60, ' '), '\n')

    if len(argv)>1: path = Path(argv[1][1:-1])
    else: path = Path('C:/Users/leona/Pictures/MonaLisas/mona.png')

    if path.is_dir():
        chdir(path)
        print(path, list(path.glob(r'*.png')))
        for arquivo in list(path.glob('*.png')):
            main(arquivo)
    else: chdir(path.parent); main(path)
